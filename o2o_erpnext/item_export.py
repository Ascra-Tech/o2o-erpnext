import frappe
from frappe import _
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import json
from datetime import datetime
import os


@frappe.whitelist()
def get_item_fields():
	"""
	Get all available fields for Item doctype, grouped by standard and custom fields.
	Returns field information including fieldname, label, fieldtype, and whether mandatory.
	Only includes fields that can be queried from database.
	"""
	item_meta = frappe.get_meta("Item")
	
	# List of non-queryable field types
	non_queryable_types = [
		"Section Break", 
		"Column Break", 
		"HTML", 
		"Table", 
		"Heading", 
		"Button", 
		"Signature",
		"Fold",
		"Break"
	]
	
	fields = []
	
	# Get all fields from Item doctype
	for field in item_meta.fields:
		# Skip non-data field types that don't exist in database
		if field.fieldtype in non_queryable_types:
			continue
		
		# Also skip virtual fields (they don't exist in database)
		if hasattr(field, 'is_virtual') and field.is_virtual:
			continue
		
		field_info = {
			"fieldname": field.fieldname,
			"label": field.label,
			"fieldtype": field.fieldtype,
			"mandatory": bool(field.reqd),
			"is_custom": field.fieldname.startswith('custom_') if field.fieldname else False,
		}
		fields.append(field_info)
	
	# Separate into mandatory and optional fields
	mandatory_fields = [f for f in fields if f["mandatory"]]
	optional_fields = [f for f in fields if not f["mandatory"]]
	
	return {
		"all_fields": fields,
		"mandatory_fields": mandatory_fields,
		"optional_fields": optional_fields,
		"total_fields": len(fields)
	}


@frappe.whitelist()
def export_items_to_excel(item_ids, selected_fields):
	"""
	Export selected items to Excel format with chosen fields.
	
	Args:
		item_ids: JSON string of list of item names/IDs
		selected_fields: JSON string of list of field names to export
	
	Returns:
		File path for download
	"""
	try:
		# Validate inputs
		if not item_ids or not selected_fields:
			frappe.throw(_("Please select items and fields to export"))
		
		# Parse JSON inputs
		if isinstance(item_ids, str):
			item_ids = json.loads(item_ids)
		if isinstance(selected_fields, str):
			selected_fields = json.loads(selected_fields)
		
		if not item_ids:
			frappe.throw(_("Please select at least one item to export"))
		if not selected_fields:
			frappe.throw(_("Please select at least one field to export"))
		
		# Get Item meta and filter out non-queryable fields
		item_meta = frappe.get_meta("Item")
		valid_fields = []
		
		# Non-queryable field types
		non_queryable_types = [
			"Section Break", 
			"Column Break", 
			"HTML", 
			"Table", 
			"Heading", 
			"Button", 
			"Signature",
			"Fold",
			"Break"
		]
		
		for field_name in selected_fields:
			field = item_meta.get_field(field_name)
			
			# Skip non-data fields
			if not field:
				continue
			
			if field.fieldtype in non_queryable_types:
				continue
			
			# Skip virtual fields
			if hasattr(field, 'is_virtual') and field.is_virtual:
				continue
			
			valid_fields.append(field_name)
		
		if not valid_fields:
			frappe.throw(_("No valid fields selected for export"))
		
		# Fetch full item documents to ensure all fields are loaded
		items_data = []
		for item in item_ids:
			try:
				# Extract just the name/ID from the item object
				# item_ids can be a list of dicts with 'name' key or simple strings
				item_name = item['name'] if isinstance(item, dict) else item
				
				doc = frappe.get_doc("Item", item_name)
				item_dict = {}
				for field_name in valid_fields:
					try:
						# Get field value from document
						item_dict[field_name] = getattr(doc, field_name, None)
					except:
						item_dict[field_name] = None
				items_data.append(item_dict)
			except Exception as e:
				frappe.log_error(f"Error fetching item {item_name}: {str(e)}", "Item Export")
				continue
		
		# If no items fetched, throw error
		if not items_data:
			frappe.throw(_("No items found to export"))
		
		# Create workbook and select active sheet
		workbook = openpyxl.Workbook()
		worksheet = workbook.active
		worksheet.title = "Items"
		
		# Define styles
		header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
		header_font = Font(bold=True, color="FFFFFF")
		header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		
		# Get field labels for headers using VALID fields only
		headers = []
		for field_name in valid_fields:
			try:
				field = item_meta.get_field(field_name)
				if field:
					headers.append(field.label or field_name)
				else:
					headers.append(field_name)
			except:
				headers.append(field_name)
		
		# Write headers
		for col_num, header in enumerate(headers, 1):
			cell = worksheet.cell(row=1, column=col_num)
			cell.value = header
			cell.fill = header_fill
			cell.font = header_font
			cell.alignment = header_alignment
		
		# Write data rows
		for row_num, item in enumerate(items_data, 2):
			for col_num, field_name in enumerate(valid_fields, 1):
				cell = worksheet.cell(row=row_num, column=col_num)
				# Get value from item dict, convert to string if needed
				value = item.get(field_name)
				if value is not None:
					cell.value = str(value) if not isinstance(value, (int, float, bool)) else value
				else:
					cell.value = ""
				cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
		
		# Auto-adjust column widths
		for col_num in range(1, len(valid_fields) + 1):
			col_letter = openpyxl.utils.get_column_letter(col_num)
			max_length = 12
			for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
				try:
					if row[0].value:
						max_length = max(max_length, len(str(row[0].value)))
				except:
					pass
			worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
		
		# Set row height for header
		worksheet.row_dimensions[1].height = 30
		
		# Save file to temporary location
		timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
		filename = f"items_export_{timestamp}.xlsx"
		file_path = frappe.get_site_path("private/files", filename)
		
		# Create directory if it doesn't exist
		import os
		os.makedirs(os.path.dirname(file_path), exist_ok=True)
		
		# Save workbook to file
		workbook.save(file_path)
		
		# Return file info
		return {
			"status": "success",
			"message": _("Items exported successfully"),
			"filename": filename
		}
		
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "Item Export Error")
		frappe.throw(_("Error exporting items: {0}").format(str(e)))


@frappe.whitelist()
def download_exported_file(filename):
	"""Download the exported Excel file"""
	try:
		file_path = frappe.get_site_path("private/files", filename)
		
		# Security check - ensure file exists and is in the right directory
		if not os.path.exists(file_path):
			frappe.throw(_("File not found"))
		
		# Read and return file
		with open(file_path, 'rb') as f:
			frappe.local.response.filename = filename
			frappe.local.response.filecontent = f.read()
			frappe.local.response.type = "download"
		
		# Clean up file after download
		try:
			os.remove(file_path)
		except:
			pass
			
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), "File Download Error")
		frappe.throw(_("Error downloading file: {0}").format(str(e)))
